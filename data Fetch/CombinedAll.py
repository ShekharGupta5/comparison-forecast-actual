from openpyxl import Workbook
from openpyxl import load_workbook

newFileWorkBook = Workbook()
Humidity = newFileWorkBook.active


CombinedDataObject = {}

for station in range(1,45):
    name = 'Station '+ str(station) + '.xlsx'
    wb = load_workbook(name)
    ws = wb.active
    dataArray = []
    for row in range(1,ws.max_row+1):
        data = ws.cell(row,2).value
        
        if(data == 'value\n'):
            #donothing
            nothing = True
        else:
            dataArray.append(data)
    
    propertyName = 'Station ' + str(station)
    CombinedDataObject[propertyName] = dataArray


#### for population combined sheeet

column = 2
for eachStation in CombinedDataObject:
    Humidity.cell(1,column).value = eachStation
    numberofrows = len(CombinedDataObject[eachStation])
    ColumnDataArray = CombinedDataObject[eachStation]

    for row in range(2,numberofrows+1):
        Humidity.cell(row,column).value = ColumnDataArray[row-1]
    
    column = column +1 
FileName = 'CombinedSheet.xlsx'
newFileWorkBook.save(FileName)
    
