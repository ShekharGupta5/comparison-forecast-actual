from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import pymongo

myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['ForecastingDB']
collection = mydb['forecastings']



fromdate = datetime.datetime(2019,12,2)
todate = datetime.datetime(2020,1,7)

parameter = 'temperature'



for station in range(1,45):
    newFileWorkBook = Workbook()
    newSheet = newFileWorkBook.active
    newSheet.title = 'Station ' + str(station)
    stationName = 'Station ' + str(station)
    docs = collection.find({'parameter':parameter, 'timestamp':{ "$gte":fromdate, "$lt":todate } ,'provider':'IITM' ,'id':stationName }).sort("timestamp",1)
    
    newSheet.cell(1,1).value = 'timestamp'
    newSheet.cell(1,2).value = 'value'
    row = 2
    for doc in docs:
        newSheet.cell(row,1).value = doc['timestamp']
        newSheet.cell(row,2).value = doc['value']
        row = row + 1
    
    newFileWorkBook.save(stationName+'.xlsx')
    
