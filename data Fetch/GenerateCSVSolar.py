from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pymongo

dbNames = ['TemperatureSolars','PercipitationSolars','RainRateSolars','Humiditys']
myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['DataFetch']

collection = mydb[dbNames[3]]

fromdate = datetime(2020,2,9)
todate = datetime(2020,2,12)





for station in range(1,45):
    newFileWorkBook = Workbook()
    newSheet = newFileWorkBook.active
    newSheet.title = 'Station ' + str(station)
    stationName = 'Station ' + str(station)
    docs = collection.find({'timestamp':{ "$gte":fromdate, "$lt":todate } ,'provider':'IITM' ,'id':stationName }).sort("timestamp",1)
    
    newSheet.cell(1,1).value = 'timestamp'
    newSheet.cell(1,2).value = 'value'
    row = 2
    for doc in docs:
        newSheet.cell(row,1).value = doc['timestamp']
        newSheet.cell(row,2).value = doc['value']
        row = row + 1
    
    newFileWorkBook.save(stationName+'.xlsx')
    
