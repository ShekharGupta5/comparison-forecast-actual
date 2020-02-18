from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pymongo,sys,os

dbNames = ['TemperatureSolars','PercipitationSolars','RainRateSolars','Humiditys']
myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['DataFetch']

collection = mydb[sys.argv[6]]
#select db from user
year = int(sys.argv[1])
fromMonth = int(sys.argv[2])
fromDay = int(sys.argv[3])
toMonth = int(sys.argv[4])
toDay = int(sys.argv[5])

fromdate = datetime(year,fromMonth,fromDay)
todate = datetime(year,toMonth,toDay)
#plus mins kuch karna h for from date either here or in pythontemepratureauotmated.py

FileWorkBook = Workbook()
Sheet = FileWorkBook.active

for station in range(1,45):
    
    col = station+1
    stationName = 'Station ' + str(station)
    docs = collection.find({'timestamp':{ "$gte":fromdate, "$lt":todate } ,'provider':'IITM' ,'id':stationName }).sort("timestamp",1)
    
    Sheet.cell(1,1).value = 'timestamp'
    Sheet.cell(1,col).value = stationName
    row = 2
    
    for doc in docs:
        Sheet.cell(row,1).value = doc['timestamp']
        Sheet.cell(row,col).value = doc['value']
        row = row + 1
    


FileWorkBook.save('CombinedSheet'+'.xlsx')
print('File is created on ',os.getcwd())
