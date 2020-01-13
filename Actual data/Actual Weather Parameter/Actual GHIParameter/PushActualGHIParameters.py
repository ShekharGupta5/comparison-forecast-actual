from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pymongo
import sys


sys.path.append('D://Software//Forecasting DB//code//src//database')

import DatabaseSelector
dbName = DatabaseSelector.DatabaseName(sys.argv[1],provider='ACTUAL')
print('The Records are inserted in ',dbName)
myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['ForecastingDB']
collection = mydb[dbName]

date = sys.argv[1]
dateParts = date.split('_')

actual_date = int(dateParts[0])

path =  'WEATHER_PARM_' + date+ '.xlsx'

wb = load_workbook(path)
ws = wb.active

owner_mapping = {
    'badwar':'mahindra',
    'barsitadesh':'arinsun',
    'ramnagar':'acme',
    'ostro':'ostro'
}
paramter_mapping = {
    'badwar':'solar',
    'barsitadesh':'solar',
    'ramnagar':'solar',
    'ostro':'wind'
}
col_mapping_sheet = {
   'barsitadesh' :3,
   'badwar':21,
   'ramnagar':15,
   'ostro':13
}
stations = ['barsitadesh','badwar','ramnagar']
for station in stations:
    allData = []
    for row in range(4,1444):

        if(actual_date <13 ):
            timestamp = ws.cell(row,1).value
            
            if(isinstance(timestamp,str)):
                timestamp = datetime.strptime(timestamp,"%m-%d-%Y %H:%M:%S")
            year = str(timestamp.year)
            hour = str(timestamp.hour)
            minute = str(timestamp.minute) 
            day = str(timestamp.month)
            month = str(timestamp.day)
        else:
            timestamp = datetime.strptime(ws.cell(row,1).value,"%d-%m-%Y %H:%M:%S")
            year = str(timestamp.year)
            hour = str(timestamp.hour)
            minute = str(timestamp.minute) 
            day = str(timestamp.day)
            month = str(timestamp.month)

        
        string_timestamp = day + '-' + month + '-' + year +' ' + hour +':'+minute
        date_timestamp = datetime.strptime(string_timestamp,"%d-%m-%Y %H:%M")
        data = ws.cell(row,col_mapping_sheet[station]).value
        owner = owner_mapping[station]
        parameter = paramter_mapping[station]
        obj = {
            "timestamp":date_timestamp,
            "value":data,
            "id":'',
            "provider":"ACTUAL",
            "owner":owner,
            "parameter":parameter
        }
        allData.append(obj)

    avg15MinBlockValue = []
    owner = owner_mapping[station]
    parameter = paramter_mapping[station]
    for index,data in enumerate(allData):
        totalValue = 0
        avgValue = 0
        if(index%15 == 0):
            for i in range(0,15):
                totalValue += allData[i+index]['value']
            avgValue = (totalValue /15)
            blockNumber = int(index/15) +1
            timestamp = allData[index]['timestamp']
            obj = {
                "timestamp":timestamp,
                "value":avgValue,
                "id":'',
                "provider":"ACTUAL",
                "owner":owner,
                "parameter":parameter
            }
            avg15MinBlockValue.append(obj)
        
        

    collection.insert_many(avg15MinBlockValue)
    print('Total ',len(avg15MinBlockValue),' documents inserted in forecasting database.')

    
    
