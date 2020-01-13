from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pymongo
import sys


myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['ForecastingDB']
collection = mydb['generations']

date = sys.argv[1]
dateParts = date.split('_')

actual_date = int(dateParts[0])

path =  'RENEWABLE_CS_' + date+ '.xlsx'

wb = load_workbook(path)
ws = wb.active

owner_mapping_sheet = {
    '2':'mahindra',
    '3':'arinsun',
    '4':'acme',
    '5':'ostro'
}
paramter_mapping_sheet = {
    '2':'solar',
    '3':'solar',
    '4':'solar',
    '5':'wind'
}
for col in range(2,6):
    allData = []
    for row in range(3,1443):
        if(actual_date <13 ):
            timestamp = ws.cell(row,1).value
            year = str(timestamp.year)
            hour = str(timestamp.hour)
            minute = str(timestamp.minute) 
            day = str(timestamp.month)
            month = str(timestamp.day)
        else:
            timestamp = datetime.strptime(ws.cell(row,1).value,"%d-%m-%Y %H:%M:%S")
            year = str(timestamp.year)
            hour = str(timestamp.hour)
            minute = str(timestamp.minute) 
            day = str(timestamp.day)
            month = str(timestamp.month)

        
        string_timestamp = day + '-' + month + '-' + year +' ' + hour +':'+minute
        date_timestamp = datetime.strptime(string_timestamp,"%d-%m-%Y %H:%M")
        data = ws.cell(row,col).value
        owner = owner_mapping_sheet[str(col)]
        parameter = paramter_mapping_sheet[str(col)]
        obj = {
            "timestamp":date_timestamp,
            "value":data,
            "id": '',
            "provider":"ACTUAL",
            "owner":owner,
            "parameter":parameter
        }
        allData.append(obj)

    avg15MinBlockValue = []
    #print('Length of the pushing array in the starting is ',len(avg15MinBlockValue))
    owner = owner_mapping_sheet[str(col)]
    parameter = paramter_mapping_sheet[str(col)]
    for index,data in enumerate(allData):
        totalValue = 0
        avgValue = 0
        if(index%15 == 0):
            for i in range(0,15):
                totalValue += allData[i+index]['value']
            avgValue = (totalValue /15)
            blockNumber = int(index/15) +1
            timestamp = allData[index]['timestamp']
            obj = {
                "timestamp":timestamp,
                "value":avgValue,
                "id":'generation',
                "provider":"ACTUAL",
                "owner":owner,
                "parameter":parameter
            }
            avg15MinBlockValue.append(obj)
        
    #print(avg15MinBlockValue)    

    collection.insert_many(avg15MinBlockValue)
    print('Total ',len(avg15MinBlockValue),' documents inserted in generations database.')
