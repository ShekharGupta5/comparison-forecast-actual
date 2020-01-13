from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pymongo
import sys

myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['ForecastingDB']
collection = mydb['winds']

date = sys.argv[1]

path = './Windspeed Extracted/'+ date +'.xlsx'

wb = load_workbook(path)
ws = wb.active

allData = []

for col in range(2,ws.max_column+1):
    station_id = ws.cell(1,col).value
    for row in range(2,ws.max_row+1):
        timestamp = ws.cell(row,1).value
        date = datetime.strptime(timestamp,"%Y-%m-%d %H:%M")
        data = ws.cell(row,col).value
        obj = {
        "timestamp":date,
        "data":data,
        "station_id":station_id
        }
        collection.insert_one(obj)

