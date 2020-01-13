from openpyxl import Workbook
from openpyxl import load_workbook
import sys

date = sys.argv[1]

fyle =   open(date +'_01/temperature.csv')

########## Getting all the first columns from the file
for row,i in zip(fyle,range(1,2)):
    if(i==1):
        firstColoumn = row.split(',')
        

newFileName = date + '.xlsx'

newFileWorkBook = Workbook()
TemperatureSheet = newFileWorkBook.active
TemperatureSheet.title = 'Temperature'

######################## Headers of the sheet

for col in range(1,len(firstColoumn)+1):
    TemperatureSheet.cell(1,col).value = firstColoumn[col-1]

##############################################

allData = []
for row in fyle:
    allData.append(row)

meaningFullData = []
for row in allData[97:193]:   ############### range is 97:193 because for the next day data is in that row in the file
    meaningFullData.append(row)

for row,d in zip(range(2,98),range(0,96)):
    dataParts = meaningFullData[d].split(',')
    TemperatureSheet.cell(row,1).value = dataParts[0]
    for col in range(2,len(dataParts)+1):
        TemperatureSheet.cell(row,col).value = float(dataParts[col-1])
        

newFileWorkBook.save(newFileName)