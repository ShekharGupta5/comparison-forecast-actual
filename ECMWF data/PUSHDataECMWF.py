from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime,timedelta
import pymongo
import sys
import owner

sys.path.append('D://Software//Forecasting DB//code//src//database')

#As we are forecasting for a day ahead and the file name is of d-1 so 
#it is causing error , so will give a date d+1 for dabasename function
dayAhead = datetime.strptime(sys.argv[1],"%Y%m%d") + timedelta(days=1)
dayAheadString = datetime.strftime(dayAhead,"%Y%m%d")
import DatabaseSelector
dbName = DatabaseSelector.DatabaseName(dayAheadString,provider='IITM')
print('The Records are inserted in ',dbName)

myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['ForecastingDB']
collection = mydb[dbName]

date = sys.argv[1]

path =  date+'ECMWF' +'.xlsx'

wb = load_workbook(path)
sheets = wb.sheetnames
allData = []

owner_obj = owner.owner_obj

for sheet in sheets:
    ws = wb[sheet]
    for col in range(2,ws.max_column+1):
        lid = ws.cell(1,col).value
        try:
            owner = owner_obj[lid]
        except:
            owner = ''
         
        for row in range(2,ws.max_row+1):
            timestamp = ws.cell(row,1).value
            date = datetime.strptime(timestamp,"%Y-%m-%d %H:%M")
            data = ws.cell(row,col).value
            obj = {
            "timestamp":date,
            "value":data,
            "id":lid,
            "provider":"ECMWF",
            "owner":owner,
            "parameter":sheet.lower()
            }
            allData.append(obj)
            #collection.insert_one(obj)

collection.insert_many(allData)
print('Total ',len(allData),' documents inserted in ',dbName,' database.')