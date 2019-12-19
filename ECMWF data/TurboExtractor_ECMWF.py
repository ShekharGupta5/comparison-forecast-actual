from openpyxl import Workbook
from openpyxl import load_workbook
import sys

date = sys.argv[1]

baseFolder = 'RE Stations Forecast Folder/' + date

fyle_humidity =   open(baseFolder +'_01/relative_humidity.csv')
fyle_wind =  open(baseFolder +'_01/windspeed_80.csv')
fyle_temperature = open(baseFolder +'_01/temperature.csv')
fyle_solar = fyle =   open(baseFolder +'_01/solarradiation.csv')

########## Getting all the first columns from the file
for row,i in zip(fyle_wind,range(1,2)):
    if(i==1):
        firstColoumn = row.split(',')
#assuming that firstcolumn of all the wind,humidity,temeprature will be same.        
fyle_wind =  open(baseFolder +'_01/windspeed_80.csv')
### doing this thing becasue loop karne se fyle_wind me changes ho rhe ha
########################################################
newFileName = date + 'ECMWF' +'.xlsx'

newFileWorkBook = Workbook()
WindSheet = newFileWorkBook.active
WindSheet.title = 'Wind'

TemperatureSheet = newFileWorkBook.create_sheet('Temperature')
HumiditySheet = newFileWorkBook.create_sheet('Humidity')
SolarSheet = newFileWorkBook.create_sheet('Solar')

######################## Headers of the sheet

for col in range(1,len(firstColoumn)+1):
    TemperatureSheet.cell(1,col).value = firstColoumn[col-1]
    HumiditySheet.cell(1,col).value = firstColoumn[col-1]
    WindSheet.cell(1,col).value = firstColoumn[col-1]
    SolarSheet.cell(1,col).value = firstColoumn[col-1]
##############################################

allDataWind = []
allDataTemperature = []
allDataHumidity = []
allDataSolar = []

for row in fyle_humidity:
    allDataHumidity.append(row)

for row in fyle_temperature:
    allDataTemperature.append(row)

for row in fyle_wind:
    allDataWind.append(row)

for row in fyle_solar:
    allDataSolar.append(row)
    
meaningFullDataWind = []
meaningFullDataHumidity = []
meaningFullDataTemperature = []
meaningFullDataSolar = []

for row in allDataWind[99:195]:   ############### range is 97:193 because for the next day data is in that row in the file
    meaningFullDataWind.append(row)

for row in allDataHumidity[99:195]:   ############### range is 97:193 because for the next day data is in that row in the file
    meaningFullDataHumidity.append(row)

for row in allDataTemperature[99:195]:   ############### range is 97:193 because for the next day data is in that row in the file
    meaningFullDataTemperature.append(row)

for row_solar in allDataSolar[99:195]:   ############### range is 100:196 because for the next day data is in that row in the file
    meaningFullDataSolar.append(row_solar)

for row,d in zip(range(2,98),range(0,96)):
    dataPartsWind = meaningFullDataWind[d].split(',')
    dataPartsTemperature = meaningFullDataTemperature[d].split(',')
    dataPartsHumidity = meaningFullDataHumidity[d].split(',')
    dataPartsSolar = meaningFullDataSolar[d].split(',')

    TemperatureSheet.cell(row,1).value = dataPartsTemperature[0]
    WindSheet.cell(row,1).value = dataPartsWind[0]
    HumiditySheet.cell(row,1).value = dataPartsHumidity[0]
    SolarSheet.cell(row,1).value = dataPartsSolar[0]
    
    for col in range(2,len(dataPartsTemperature)+1):
        TemperatureSheet.cell(row,col).value = float(dataPartsTemperature[col-1])

    for col in range(2,len(dataPartsWind)+1):
        WindSheet.cell(row,col).value = float(dataPartsWind[col-1])

    for col in range(2,len(dataPartsHumidity)+1):
        HumiditySheet.cell(row,col).value = float(dataPartsHumidity[col-1])

    for col in range(2,len(dataPartsSolar)+1):
        SolarSheet.cell(row,col).value = float(dataPartsSolar[col-1])
        

newFileWorkBook.save(newFileName)