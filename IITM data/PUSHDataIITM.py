from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pymongo
import sys
import owner

myclient = pymongo.MongoClient("mongodb://127.0.0.1:27017/")
mydb = myclient['ForecastingDB']
collection = mydb['forecastings']

date = sys.argv[1]

path =  date+'IITM' +'.xlsx'

wb = load_workbook(path)
sheets = wb.sheetnames
allData = []

owner_obj = owner.owner_obj

for sheet in sheets:
    ws = wb[sheet]
    for col in range(2,ws.max_column+1):
        stn = ws.cell(1,col).value
        stn_id = stn.split(' ')[1]
        try:
            owner = owner_obj[stn_id]
        except:
            owner = ''
         
        for row in range(2,ws.max_row+1):
            timestamp = ws.cell(row,1).value
            date = datetime.strptime(timestamp,"%d-%m-%Y %H:%M")
            data = ws.cell(row,col).value
            obj = {
            "timestamp":date,
            "value":data,
            "id":stn,
            "provider":"IITM",
            "owner":owner,
            "parameter":sheet.lower()
            }
            allData.append(obj)
            #collection.insert_one(obj)

collection.insert_many(allData)
print('Total ',len(allData),' documents inserted in forecasting database.')