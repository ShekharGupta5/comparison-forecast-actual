from openpyxl import Workbook
from openpyxl import load_workbook
import sys
date = sys.argv[1]

base = 'wrf_other.' + date +'00/output_Pstn'

newFileName = date + '.xlsx'

newFileWorkBook = Workbook()
WindSheet = newFileWorkBook.active
WindSheet.title = 'Wind'

TemperatureSheet = newFileWorkBook.create_sheet('Temperature')
HumiditySheet = newFileWorkBook.create_sheet('Humidity')

##   Headers of the sheet
WindSheet.cell(1,1).value = 'TIME'
TemperatureSheet.cell(1,1).value = 'TIME'
HumiditySheet.cell(1,1).value = 'TIME'

for col in range(2,46):
    WindSheet.cell(1,col).value = 'Station '+ str(col-1)
    TemperatureSheet.cell(1,col).value = 'Station '+ str(col-1)
    HumiditySheet.cell(1,col).value = 'Station '+ str(col-1)

############################

for station in range(1,45):
    allData = []
    if(station<10):
        name = base+ '0' + str(station) + '.txt'
    else:
        name = base+ str(station) + '.txt'
    

    fyle = open(name)
    
    #print('File ',name,' is open and working...')
    
    for row in fyle:
        allData.append(row)


    #### For loop for inserting data into sheet
    for row,d in zip(range(2,98),range(95,191)):
        dataParts = allData[d].split(' ')
        WindSheet.cell(row,1).value = dataParts[1]
        WindSheet.cell(row,station+1).value = dataParts[2]
        TemperatureSheet.cell(row,1).value = dataParts[1]
        HumiditySheet.cell(row,1).value = dataParts[1]
        TemperatureSheet.cell(row,station+1).value = dataParts[3]
        HumiditySheet.cell(row,station+1).value = dataParts[4]
    




newFileWorkBook.save(newFileName)
    