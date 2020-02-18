from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime,timedelta
import sys

date = sys.argv[1]

base = 'RE Stations Forecast Folder/Other/wrf_other.' + date +'00/output_Pstn'

newFileName = date + 'IITM' +'.xlsx'

newFileWorkBook = Workbook()
WindSheet = newFileWorkBook.active
WindSheet.title = 'Wind'

TemperatureSheet = newFileWorkBook.create_sheet('Temperature')
HumiditySheet = newFileWorkBook.create_sheet('Humidity')

##   Headers of the sheet
WindSheet.cell(1,1).value = 'TIME'
TemperatureSheet.cell(1,1).value = 'TIME'
HumiditySheet.cell(1,1).value = 'TIME'

for col in range(2,371):
    WindSheet.cell(1,col).value = 'Station '+ str(col-1)
    TemperatureSheet.cell(1,col).value = 'Station '+ str(col-1)
    HumiditySheet.cell(1,col).value = 'Station '+ str(col-1)
#first col is reserved for timestamp 
############################


for station in range(1,370):
    allData = []
    if(station<10):
        name = base+ '0' + str(station) + '.csv'
    else:
        name = base+ str(station) + '.csv'
    
    
    fyle = open(name)
    
    #print('File ',name,' is open and working...')
    
    for row in fyle:
        allData.append(row)


    #### For loop for inserting data into sheet
    for row,d in zip(range(2,98),range(73,169)):
        # print('  THe value for which it is breaking ',station)
        dataParts = allData[d].split(' ')
        dateValue = datetime.strptime(dataParts[0] +' '+ dataParts[1],"%d-%m-%Y %H:%M")
        dateValue = dateValue + timedelta(hours=5) + timedelta(minutes=30)
        dateValue = datetime.strftime(dateValue,"%d-%m-%Y %H:%M")
        WindSheet.cell(row,1).value = dateValue
        TemperatureSheet.cell(row,1).value = dateValue
        HumiditySheet.cell(row,1).value = dateValue
        WindSheet.cell(row,station+1).value = float(dataParts[2])
        TemperatureSheet.cell(row,station+1).value = float(dataParts[3])
        try:
            humidiyValue = float(dataParts[4])
        except:
            humidiyValue = 0
        HumiditySheet.cell(row,station+1).value = humidiyValue
    




newFileWorkBook.save(newFileName)
    

    
